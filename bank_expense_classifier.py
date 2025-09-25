import sys
import logging
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, NamedStyle
import json
import os

# === LOGGING SETUP ===
LOG_PATH = os.path.join(os.path.dirname(__file__), "error.log")
logging.basicConfig(filename=LOG_PATH, level=logging.ERROR, format='%(asctime)s %(levelname)s: %(message)s')

# === LOAD STATIC VALUES ===
STATIC_PATH = os.path.join(os.path.dirname(__file__), "static_values.json")
with open(STATIC_PATH, "r") as f:
    STATIC = json.load(f)

# === CLASSIFICATION & SUMMARIZATION ===
def classify_fs_line(details):
    details_lower = details.lower()
    if any(keyword in details_lower for keyword in STATIC["salary_keywords"]):
        return "Salary"
    elif any(keyword in details_lower for keyword in STATIC["utility_keywords"]):
        return "Utility"
    elif any(keyword in details_lower for keyword in STATIC["rent_keywords"]):
        return "Rent"
    elif any(keyword in details_lower for keyword in STATIC["ingredient_keywords"]):
        return "Ingredient"
    else:
        return "Other"

def summarize_item(details):
    details_lower = details.lower()
    if any(keyword in details_lower for keyword in STATIC["utility_keywords"]):
        return "Utilities"
    elif any(keyword in details_lower for keyword in STATIC["rent_keywords"]):
        return "Rent"
    elif any(name.lower() in details_lower for name in STATIC["names"]):
        matched_name = next(name for name in STATIC["names"] if name.lower() in details_lower)
        return f"Salary - {matched_name}"
    elif any(keyword in details_lower for keyword in STATIC["milk_keywords"]):
        return "Milk Purchase"
    elif any(keyword in details_lower for keyword in STATIC["coffee_keywords"]):
        return "Coffee Purchase"
    else:
        words = details.split()
        return " ".join(words[:5]) if words else "Unknown"



# === CONSTANTS ===
HEADER_SEARCH_ROWS = 20  # Number of rows to search for header
LOG_FILE_NAME = "error.log"

# === FILE PROCESSING ===
def process_file(filepath, currency):
    if not os.path.isfile(filepath):
        raise FileNotFoundError(f"File not found: {filepath}")
    df = pd.read_excel(filepath, header=None)
    header_row = None
    for i in range(min(HEADER_SEARCH_ROWS, len(df))):
        if df.iloc[i].astype(str).str.contains("Date", case=False, na=False).any():
            header_row = i
            break
    if header_row is None:
        raise ValueError("Could not find header row with 'Date'")

    df = pd.read_excel(filepath, header=header_row)
    df = df.rename(columns=lambda x: str(x).strip().lower())

    def find_column(col_key):
        matches = [col for col in df.columns if STATIC["column_names"][col_key].lower() in col.lower()]
        if not matches:
            raise ValueError(f"Column '{STATIC['column_names'][col_key]}' not found in file.")
        return matches[0]

    date_col = find_column("date")
    details_col = find_column("details")
    money_out_col = find_column("money_out")

    # Filter out rows where "Money Out" is empty or zero
    df = df[df[money_out_col].notna() & (df[money_out_col] != 0)]

    output_rows = []
    for _, row in df.iterrows():
        try:
            date = pd.to_datetime(row[date_col])
            datecode = f"{date.month}{date.year}"
            details = str(row[details_col])
            fs_line = classify_fs_line(details)
            item = summarize_item(details)
            expense_usd = row[money_out_col] if currency == "USD" else 0
            expense_khr = row[money_out_col] if currency == "KHR" else 0

            output_rows.append({
                "Datecode": datecode,
                "Date": date.strftime("%Y-%m-%d"),
                "FS Line": fs_line,
                "Item": item,
                "Expense USD": expense_usd,
                "Expense KHR": expense_khr
            })
        except Exception as e:
            logging.error(f"Error processing row: {e}")

    return pd.DataFrame(output_rows)

# === GUI ===
def run_gui():
    def load_khr():
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path:
            khr_path.set(path)

    def load_usd():
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path:
            usd_path.set(path)

    def generate_report():
        try:
            df_khr = process_file(khr_path.get(), "KHR") if khr_path.get() else pd.DataFrame()
            df_usd = process_file(usd_path.get(), "USD") if usd_path.get() else pd.DataFrame()
            final_df = pd.concat([df_khr, df_usd], ignore_index=True)

            final_df["Date"] = pd.to_datetime(final_df["Date"])
            final_df = final_df.sort_values(by="Date")
            final_df["Date"] = final_df["Date"].dt.strftime("%Y-%m-%d")

            # Clean and ensure "Expense USD" and "Expense KHR" columns are numeric
            final_df["Expense USD"] = (
                final_df["Expense USD"]
                .astype(str)  # Convert to string to handle any non-numeric values
                .str.replace(",", "", regex=True)  # Remove commas
                .str.strip()  # Remove leading/trailing spaces
            )
            final_df["Expense KHR"] = (
                final_df["Expense KHR"]
                .astype(str)
                .str.replace(",", "", regex=True)
                .str.strip()
            )
            final_df["Expense USD"] = pd.to_numeric(final_df["Expense USD"], errors="coerce")
            final_df["Expense KHR"] = pd.to_numeric(final_df["Expense KHR"], errors="coerce")

            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if save_path:
                # Save the DataFrame to Excel
                final_df.to_excel(save_path, index=False)

                # Apply accounting number format using openpyxl
                wb = load_workbook(save_path)
                ws = wb.active

                # Define accounting style
                accounting_style = NamedStyle(name="AccountingStyle", number_format="#,##0.00_);(#,##0.00)")

                # Register the style if not already registered
                if "AccountingStyle" not in wb.named_styles:
                    wb.add_named_style(accounting_style)


                # Dynamically find column indexes for 'Expense USD' and 'Expense KHR'
                header = [cell.value for cell in ws[1]]
                try:
                    usd_idx = header.index("Expense USD") + 1  # openpyxl is 1-indexed
                    khr_idx = header.index("Expense KHR") + 1
                except ValueError:
                    raise Exception("Could not find 'Expense USD' or 'Expense KHR' columns in the output file.")

                for row in ws.iter_rows(min_row=2, min_col=usd_idx, max_col=usd_idx):
                    for cell in row:
                        cell.style = accounting_style

                for row in ws.iter_rows(min_row=2, min_col=khr_idx, max_col=khr_idx):
                    for cell in row:
                        cell.style = accounting_style

                # Save the workbook with formatting
                wb.save(save_path)

                messagebox.showinfo("Success", f"Expense report saved to:\n{save_path}")
            else:
                messagebox.showinfo("Cancelled", "Report generation cancelled.")
        except Exception as e:
            logging.error(f"Error in report generation: {e}")
            messagebox.showerror("Error", f"{str(e)}\nSee error.log for details.")

    root = tk.Tk()
    root.title("Expense Report Generator")

    khr_path = tk.StringVar()
    usd_path = tk.StringVar()

    tk.Label(root, text="KHR Statement:").grid(row=0, column=0, padx=10, pady=5, sticky="e")
    tk.Entry(root, textvariable=khr_path, width=50).grid(row=0, column=1)
    tk.Button(root, text="Browse", command=load_khr).grid(row=0, column=2)

    tk.Label(root, text="USD Statement:").grid(row=1, column=0, padx=10, pady=5, sticky="e")
    tk.Entry(root, textvariable=usd_path, width=50).grid(row=1, column=1)
    tk.Button(root, text="Browse", command=load_usd).grid(row=1, column=2)

    tk.Button(root, text="Generate Report", command=generate_report, bg="green", fg="white").grid(row=2, column=1, pady=20)

    root.mainloop()

# === RUN ===
if __name__ == "__main__":
    run_gui()
